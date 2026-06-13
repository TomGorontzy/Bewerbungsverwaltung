from __future__ import annotations

import warnings
from copy import copy, deepcopy
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .constants import (
    ARCHIVE_STATUS,
    DEFAULT_NAECHSTER_SCHRITT_CATALOG,
    FIELD_ORDER,
    FOLLOW_UP_STATUS,
    LOOKUP_COLUMNS,
    SHEET_LOOKUPS,
    SHEET_OVERVIEW,
)
from .models import ApplicationRecord, LookupValues
from .utils import as_date, as_str, today


DATE_FIELDS = {"datum_bewerbung", "status_datum", "letzte_kontaktaufnahme", "erinnerungsdatum"}
DATE_NUMBER_FORMAT = "dd.mm.yyyy"
SHEET_TODAY = "Heute erledigen"
SHEET_THIS_WEEK = "Diese Woche erledigen"
SUMMARY_MAX_ROW = 100
SOURCE_DATA_START_ROW = 2
LOOKUP_FIELD_TO_OVERVIEW_COLUMN = {
    "unterlagen": FIELD_ORDER["unterlagen"],
    "art_bewerbung": FIELD_ORDER["art_bewerbung"],
    "status": FIELD_ORDER["status"],
    "art_kontaktaufnahme": FIELD_ORDER["art_kontaktaufnahme"],
    "naechster_schritt": FIELD_ORDER["naechster_schritt"],
    "prioritaet": FIELD_ORDER["prioritaet"],
    "endergebnis": FIELD_ORDER["endergebnis"],
    "quelle": FIELD_ORDER["quelle"],
}


class ExcelRepository:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path

    @staticmethod
    def _is_formula(value: Any) -> bool:
        return isinstance(value, str) and value.startswith("=")

    @staticmethod
    def _build_bewerbungs_id(row: int, datum_bewerbung) -> str:
        if datum_bewerbung is None:
            return ""
        return f"BEW-{datum_bewerbung.year}-{row - 1:03d}"

    @staticmethod
    def _load_workbook(*args, **kwargs):
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Data Validation extension is not supported and will be removed",
                category=UserWarning,
                module=r"openpyxl\..*",
            )
            return load_workbook(*args, **kwargs)

    def load_lookups(self) -> LookupValues:
        wb = self._load_workbook(self.workbook_path, data_only=True)
        ws = wb[SHEET_LOOKUPS]
        values: dict[str, list[str]] = {}

        for key, col in LOOKUP_COLUMNS.items():
            items: list[str] = []
            last_row = self._find_last_lookup_row(ws, col)
            for row in range(2, last_row + 1):
                value = as_str(ws[f"{col}{row}"].value)
                if value and value not in items:
                    items.append(value)
            values[key] = items

        wb.close()
        return LookupValues(**values)

    def load_records(self) -> list[ApplicationRecord]:
        wb = self._load_workbook(self.workbook_path, data_only=True)
        ws = wb[SHEET_OVERVIEW]
        records: list[ApplicationRecord] = []

        for row in range(2, ws.max_row + 1):
            unternehmen = as_str(ws.cell(row=row, column=FIELD_ORDER["unternehmen"]).value)
            position = as_str(ws.cell(row=row, column=FIELD_ORDER["position"]).value)
            if not unternehmen and not position:
                continue

            values: dict[str, Any] = {"row": row}
            for field, col in FIELD_ORDER.items():
                raw = ws.cell(row=row, column=col).value
                if field in {"datum_bewerbung", "status_datum", "letzte_kontaktaufnahme", "erinnerungsdatum"}:
                    values[field] = as_date(raw)
                else:
                    values[field] = as_str(raw)

            if not values["bewerbungs_id"]:
                values["bewerbungs_id"] = self._build_bewerbungs_id(row, values.get("datum_bewerbung"))

            if not values["erinnerungsdatum"] and values["status"] in FOLLOW_UP_STATUS and values["status_datum"] is not None:
                values["erinnerungsdatum"] = values["status_datum"] + timedelta(days=14)

            records.append(ApplicationRecord(**values))

        wb.close()
        return records

    def ensure_workbook_health(self) -> None:
        """Prüft beim Start, ob Restauration nötig ist, und führt diese durch."""
        wb = self._load_workbook(self.workbook_path, data_only=False)
        catalog_changed = self._ensure_default_naechster_schritt_catalog(wb)
        needs_repair = (
            catalog_changed
            or
            self._needs_data_validation_repair(wb)
            or self._needs_cf_repair(wb)
            or self._needs_summary_formula_repair(wb)
        )
        if needs_repair:
            self._repair_summary_sheet_formulas(wb)
            wb.save(self.workbook_path)
        wb.close()

    def _needs_data_validation_repair(self, wb) -> bool:
        if SHEET_OVERVIEW not in wb.sheetnames:
            return False
        dvs = getattr(wb[SHEET_OVERVIEW].data_validations, 'dataValidation', [])
        return len(dvs) < len(LOOKUP_COLUMNS)

    def _needs_cf_repair(self, wb) -> bool:
        if SHEET_OVERVIEW not in wb.sheetnames:
            return False
        cf = wb[SHEET_OVERVIEW].conditional_formatting
        return len(list(cf)) == 0

    def _needs_summary_formula_repair(self, wb) -> bool:
        for sheet_name in (SHEET_TODAY, SHEET_THIS_WEEK):
            if sheet_name not in wb.sheetnames:
                return True
            cell_a2 = wb[sheet_name]['A2'].value
            if not cell_a2 or not str(cell_a2).startswith('=IFERROR'):
                return True
        return False

    def add_record(self, payload: dict[str, Any]) -> None:
        wb = self._load_workbook(self.workbook_path, data_only=False)
        ws = wb[SHEET_OVERVIEW]

        row = ws.max_row + 1
        self._write_payload(ws, row, payload)
        self._apply_default_formulas(ws, row)
        self._ensure_default_naechster_schritt_catalog(wb)
        self._repair_summary_sheet_formulas(wb)
        self._repair_overview_data_validations(wb)
        self._repair_overview_conditional_formatting(wb)

        wb.save(self.workbook_path)
        wb.close()

    def update_record(self, row: int, updates: dict[str, Any]) -> None:
        wb = self._load_workbook(self.workbook_path, data_only=False)
        ws = wb[SHEET_OVERVIEW]

        self._write_payload(ws, row, updates, partial=True)
        self._apply_default_formulas(ws, row)
        self._ensure_default_naechster_schritt_catalog(wb)
        self._repair_summary_sheet_formulas(wb)
        self._repair_overview_data_validations(wb)
        self._repair_overview_conditional_formatting(wb)

        wb.save(self.workbook_path)
        wb.close()

    def save_naechster_schritt_catalog(self, items: list[str]) -> None:
        cleaned: list[str] = []
        seen: set[str] = set()
        for item in items:
            value = as_str(item)
            if not value:
                continue
            marker = value.casefold()
            if marker in seen:
                continue
            cleaned.append(value)
            seen.add(marker)

        if not cleaned:
            raise ValueError("Der Katalog darf nicht leer sein.")

        wb = self._load_workbook(self.workbook_path, data_only=False)
        if SHEET_LOOKUPS not in wb.sheetnames or "naechster_schritt" not in LOOKUP_COLUMNS:
            wb.close()
            raise KeyError("Hilfstabellen oder Lookup-Spalte für 'Nächster Schritt' fehlt.")

        ws = wb[SHEET_LOOKUPS]
        column_letter = LOOKUP_COLUMNS["naechster_schritt"]

        for row in range(2, max(ws.max_row, SUMMARY_MAX_ROW) + 1):
            ws[f"{column_letter}{row}"] = None

        for idx, value in enumerate(cleaned, start=2):
            ws[f"{column_letter}{idx}"] = value

        self._repair_overview_data_validations(wb)
        wb.save(self.workbook_path)
        wb.close()

    def categorize_records(self, records: list[ApplicationRecord]) -> tuple[list[ApplicationRecord], list[ApplicationRecord]]:
        follow_up: list[ApplicationRecord] = []
        archive: list[ApplicationRecord] = []

        now = today()
        for record in records:
            archived = bool(record.endergebnis) or record.status in ARCHIVE_STATUS
            if archived:
                archive.append(record)
                continue

            due_by_reminder = record.erinnerungsdatum is not None and record.erinnerungsdatum <= now
            due_by_status_age = (
                record.status in FOLLOW_UP_STATUS
                and record.status_datum is not None
                and (now - record.status_datum).days >= 14
            )
            flagged_today = record.heute_erledigen.lower() == "ja"

            if due_by_reminder or due_by_status_age or flagged_today:
                follow_up.append(record)

        follow_up.sort(key=lambda x: (x.erinnerungsdatum or date.max, x.prioritaet != "Hoch"))
        archive.sort(key=lambda x: (x.status_datum or date.min), reverse=True)
        return follow_up, archive

    def _write_payload(self, ws, row: int, payload: dict[str, Any], partial: bool = False) -> None:
        for field, value in payload.items():
            if field not in FIELD_ORDER:
                continue
            col = FIELD_ORDER[field]
            if value is None and partial:
                continue
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if field in DATE_FIELDS and value is not None:
                cell.number_format = DATE_NUMBER_FORMAT

    @staticmethod
    def _repair_summary_sheet_formulas(wb) -> None:
        if getattr(wb, "calculation", None) is not None:
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        today_criteria = '(Bewerbungsübersicht!$V$2:$V$100="ja")'
        week_criteria = (
            '(Bewerbungsübersicht!$R$2:$R$100<>"")*'
            '(Bewerbungsübersicht!$R$2:$R$100>TODAY())*'
            '(Bewerbungsübersicht!$R$2:$R$100<=TODAY()+7)*'
            '(Bewerbungsübersicht!$T$2:$T$100="")'
        )
        ExcelRepository._rebuild_summary_sheet(wb, SHEET_TODAY, today_criteria)
        ExcelRepository._rebuild_summary_sheet(wb, SHEET_THIS_WEEK, week_criteria)
        ExcelRepository._repair_overview_data_validations(wb)
        ExcelRepository._repair_overview_conditional_formatting(wb)

    @staticmethod
    def _repair_overview_data_validations(wb) -> None:
        if SHEET_OVERVIEW not in wb.sheetnames or SHEET_LOOKUPS not in wb.sheetnames:
            return

        overview_ws = wb[SHEET_OVERVIEW]
        lookups_ws = wb[SHEET_LOOKUPS]
        overview_ws.data_validations.dataValidation = []

        for lookup_field, lookup_column in LOOKUP_COLUMNS.items():
            overview_column = LOOKUP_FIELD_TO_OVERVIEW_COLUMN.get(lookup_field)
            if overview_column is None:
                continue

            last_row = ExcelRepository._find_last_lookup_row(lookups_ws, lookup_column)
            if last_row < 2:
                continue

            formula = f"='{SHEET_LOOKUPS}'!${lookup_column}$2:${lookup_column}${last_row}"
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            validation.prompt = "Bitte einen Wert aus der Liste auswählen."
            validation.promptTitle = "Auswahlliste"
            overview_ws.add_data_validation(validation)

            column_letter = get_column_letter(overview_column)
            validation.add(f"{column_letter}{SOURCE_DATA_START_ROW}:{column_letter}{SUMMARY_MAX_ROW}")

    @staticmethod
    def _find_last_lookup_row(ws, column_letter: str) -> int:
        for row in range(ws.max_row, 1, -1):
            if as_str(ws[f"{column_letter}{row}"].value):
                return row
        return 1

    @staticmethod
    def _ensure_default_naechster_schritt_catalog(wb) -> bool:
        if SHEET_LOOKUPS not in wb.sheetnames or "naechster_schritt" not in LOOKUP_COLUMNS:
            return False

        ws = wb[SHEET_LOOKUPS]
        column_letter = LOOKUP_COLUMNS["naechster_schritt"]
        existing_values: list[str] = []

        for row in range(2, max(ws.max_row, 2) + 1):
            value = as_str(ws[f"{column_letter}{row}"].value)
            if value:
                existing_values.append(value)

        if existing_values:
            return False

        last_row = ExcelRepository._find_last_lookup_row(ws, column_letter)
        next_row = max(2, last_row + 1)
        changed = False

        for item in DEFAULT_NAECHSTER_SCHRITT_CATALOG:
            ws[f"{column_letter}{next_row}"] = item
            next_row += 1
            changed = True

        return changed

    @staticmethod
    def _repair_overview_conditional_formatting(wb) -> None:
        if SHEET_OVERVIEW not in wb.sheetnames or 'Erläuterungen' not in wb.sheetnames:
            return

        overview_ws = wb[SHEET_OVERVIEW]
        template_ws = wb['Erläuterungen']
        overview_ws.conditional_formatting._cf_rules.clear()

        for cf_range in list(template_ws.conditional_formatting):
            for rule in template_ws.conditional_formatting[cf_range]:
                cloned_rule = Rule(
                    type=rule.type,
                    dxf=deepcopy(rule.dxf),
                    stopIfTrue=rule.stopIfTrue,
                    aboveAverage=getattr(rule, 'aboveAverage', None),
                    percent=getattr(rule, 'percent', None),
                    bottom=getattr(rule, 'bottom', None),
                    operator=getattr(rule, 'operator', None),
                    text=getattr(rule, 'text', None),
                    timePeriod=getattr(rule, 'timePeriod', None),
                    rank=getattr(rule, 'rank', None),
                    stdDev=getattr(rule, 'stdDev', None),
                    equalAverage=getattr(rule, 'equalAverage', None),
                    formula=list(getattr(rule, 'formula', []) or []),
                )
                overview_ws.conditional_formatting.add(f"A2:V100", cloned_rule)

    @staticmethod
    def _rebuild_summary_sheet(wb, sheet_name: str, criteria: str) -> None:
        overview = wb[SHEET_OVERVIEW]
        sheet_index = wb.sheetnames.index(sheet_name) if sheet_name in wb.sheetnames else len(wb.sheetnames)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        ws = wb.create_sheet(title=sheet_name, index=sheet_index)

        for col in range(1, len(FIELD_ORDER) + 1):
            source_cell = overview.cell(row=1, column=col)
            target_cell = ws.cell(row=1, column=col, value=source_cell.value)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.protection = copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format
            column_letter = source_cell.column_letter
            ws.column_dimensions[column_letter].width = overview.column_dimensions[column_letter].width

        for row in range(2, SUMMARY_MAX_ROW + 1):
            for col in range(1, len(FIELD_ORDER) + 1):
                source_column_letter = overview.cell(row=1, column=col).column_letter
                formula = ExcelRepository._build_summary_cell_formula(source_column_letter, criteria, row)
                cell = ws.cell(row=row, column=col, value=formula)
                source_number_format = overview.cell(row=2, column=col).number_format
                if source_number_format:
                    cell.number_format = source_number_format
                field = list(FIELD_ORDER.keys())[col - 1]
                if field in DATE_FIELDS:
                    cell.number_format = DATE_NUMBER_FORMAT

        ws.freeze_panes = "A2"

    @staticmethod
    def _build_summary_cell_formula(source_column_letter: str, criteria: str, target_row: int) -> str:
        row_selector = f'ROWS($A$2:A{target_row})'
        value_expr = (
            f'INDEX(_xlfn._xlws.FILTER(Bewerbungsübersicht!{source_column_letter}$2:{source_column_letter}$100,{criteria}),'
            f'{row_selector})'
        )
        return (
            f'=IFERROR(IF({value_expr}=0,"",{value_expr}),"")'
        )

    @staticmethod
    def _apply_default_formulas(ws, row: int) -> None:
        ws.cell(row=row, column=FIELD_ORDER["bewerbungs_id"]).value = (
            f'=IF(B{row}="","","BEW-"&YEAR(J{row})&"-"&TEXT(ROW()-1,"000"))'
        )
        erinnerungsdatum_cell = ws.cell(row=row, column=FIELD_ORDER["erinnerungsdatum"])
        erinnerungsdatum_cell.value = (
            f'=IF(OR(L{row}="Bewerbung versendet",L{row}="Eingangsbestätigung",L{row}="Rückmeldung ausstehend"),M{row}+14,"")'
        )
        erinnerungsdatum_cell.number_format = DATE_NUMBER_FORMAT
        ws.cell(row=row, column=FIELD_ORDER["heute_erledigen"]).value = (
            f'=IF(AND(R{row}<>"",R{row}<=TODAY(),T{row}=""),"ja","")'
        )
